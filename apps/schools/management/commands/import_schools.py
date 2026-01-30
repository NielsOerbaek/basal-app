from collections import defaultdict

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from apps.schools.models import Person, PersonRole, School


class Command(BaseCommand):
    help = "Importer skoler fra Excel-fil"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Sti til Excel-fil",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Vis hvad der ville blive importeret uden at gemme",
        )
        parser.add_argument(
            "--sheet",
            type=int,
            default=0,
            help="Sheet index (0-baseret, standard: 0)",
        )
        parser.add_argument(
            "--start-row",
            type=int,
            default=5,
            help="Første række med data (standard: 5, efter header)",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        dry_run = options["dry_run"]
        sheet_index = options["sheet"]
        start_row = options["start_row"]

        self.stdout.write(f"Læser fil: {file_path}")

        try:
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.worksheets[sheet_index]
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Kunne ikke åbne fil: {e}"))
            return

        # Parse rows and group by school
        schools_data = defaultdict(list)

        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # Skip empty rows
            if not row[0] and not row[1]:
                continue

            kommune = self.clean_value(row[0])
            skole = self.clean_value(row[1])
            navn = self.clean_value(row[2])
            titel = self.clean_value(row[3])
            telefon = self.clean_value(row[4])
            email = self.clean_value(row[5])

            if not skole:
                continue

            school_key = (kommune, skole)

            # Initialize school entry if not exists
            if school_key not in schools_data:
                schools_data[school_key] = []

            # Only add person if name is valid (not empty or just "?")
            if navn and navn not in ("?", "-", ""):
                schools_data[school_key].append(
                    {
                        "navn": navn,
                        "titel": titel,
                        "telefon": str(telefon) if telefon else "",
                        "email": email,
                    }
                )

        wb.close()

        if dry_run:
            self.stdout.write(self.style.WARNING("\n=== DRY RUN - Ingen data gemmes ===\n"))

        schools_created = 0
        schools_existing = 0
        people_created = 0

        for (kommune, skole_navn), people in schools_data.items():
            if dry_run:
                self.stdout.write(f"\nSkole: {skole_navn} ({kommune})")
                for person in people:
                    self.stdout.write(f'  - {person["navn"]} ({person["titel"]})')
                    if person["email"]:
                        self.stdout.write(f'    Email: {person["email"]}')
                    if person["telefon"]:
                        self.stdout.write(f'    Tlf: {person["telefon"]}')
            else:
                # Create or get school
                school, created = School.objects.get_or_create(
                    name=skole_navn,
                    kommune=kommune,
                    defaults={
                        "adresse": "",
                    },
                )

                if created:
                    schools_created += 1
                    self.stdout.write(self.style.SUCCESS(f"Oprettet skole: {skole_navn} ({kommune})"))
                else:
                    schools_existing += 1
                    self.stdout.write(f"Skole findes allerede: {skole_navn} ({kommune})")

                # Create people
                for person in people:
                    if not person["navn"]:
                        continue

                    # Map titel to PersonRole
                    role = self.map_role(person["titel"])

                    # Check if person already exists
                    existing = Person.objects.filter(
                        school=school,
                        name=person["navn"],
                    ).first()

                    if existing:
                        self.stdout.write(f'  Person findes allerede: {person["navn"]}')
                        continue

                    Person.objects.create(
                        school=school,
                        name=person["navn"],
                        role=role,
                        role_other=person["titel"] if role == PersonRole.OTHER else "",
                        phone=person["telefon"],
                        email=person["email"],
                    )
                    people_created += 1
                    self.stdout.write(f'  Oprettet person: {person["navn"]}')

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING(f"Ville oprette {len(schools_data)} skoler"))
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f"Import færdig: {schools_created} skoler oprettet, "
                    f"{schools_existing} fandtes allerede, {people_created} personer oprettet"
                )
            )

    def clean_value(self, value):
        """Clean and normalize a cell value."""
        if value is None:
            return ""
        if isinstance(value, str):
            # Remove zero-width characters and normalize whitespace
            value = value.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "")
            return value.strip()
        return value

    def map_role(self, titel):
        """Map a title string to a PersonRole."""
        if not titel:
            return PersonRole.OTHER

        titel_lower = titel.lower()

        if "koordinator" in titel_lower:
            return PersonRole.KOORDINATOR
        elif "skoleleder" in titel_lower and "udskoling" not in titel_lower:
            return PersonRole.SKOLELEDER
        elif "udskoling" in titel_lower:
            return PersonRole.UDSKOLINGSLEDER
        else:
            return PersonRole.OTHER
