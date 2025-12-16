import openpyxl
from django.http import HttpResponse


def export_queryset_to_excel(queryset, fields, filename):
    """
    Generic Excel export utility.

    Args:
        queryset: Django queryset to export
        fields: List of (field_name, header_label) tuples
        filename: Output filename (without extension)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.capitalize()

    # Headers
    for col, (_, header) in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for row_num, obj in enumerate(queryset, start=2):
        for col, (field, _) in enumerate(fields, start=1):
            # Handle nested attributes (e.g., 'school__name')
            if '__' in field:
                parts = field.split('__')
                value = obj
                for part in parts:
                    value = getattr(value, part, '')
                    if value is None:
                        value = ''
                        break
            else:
                value = getattr(obj, field, '')

            # Handle callable properties
            if callable(value):
                value = value()

            # Convert to string for Excel
            if value is not None:
                ws.cell(row=row_num, column=col, value=str(value))
            else:
                ws.cell(row=row_num, column=col, value='')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
